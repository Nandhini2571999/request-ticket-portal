import os
import sys
import shutil
import pandas as pd 
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from datetime import date, timedelta,datetime

class ConvertFile:

    def __init__(self):
        self.today = date.today()
        self.BASE_DIR = r'C:\Users\nandh\MyProject\PythonProjects\TASCOutsourcing\Emp_IN_OUT\{}'.format(self.today.strftime("%Y-%m-%d"))
        file_exists =  [i for i in os.listdir(self.BASE_DIR) if i.endswith('.xlsx')]
        if file_exists:
            self.input_filepath = os.path.join(self.BASE_DIR,file_exists[0])
        else:
            print(f"No InputFile placed")
            sys.exit()

    def get_actual_working_days(self):
        today = date.today()
        if today.month == 1:
            last_month = 12
            year = today.year - 1
        else:
            last_month = today.month - 1
            year = today.year

        start_date = date(year, last_month, 20)
        end_date = date(today.year, today.month, 18)

        # Calculate actual days
        no_of_actual_days = (end_date - start_date).days + 1  # +1 to include both dates
        return start_date, no_of_actual_days

    def calculate_leave_end_date(self, start_date_str, leave_days):

        start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
        end_date = start_date + timedelta(days=leave_days - 1)
        return end_date

    def generate_outputfile(self):
        start_date, no_of_actual_days = self.get_actual_working_days()
        df = pd.read_excel(self.input_filepath, skiprows=1).fillna('')
        content_list = []
        for index, row in df.iterrows():
            if row.get('Instashop ID', '') == '':
                break
            fields = {}
            fields['SAP ID'] = int(row.get('Instashop ID', ''))
            fields['TM Name'] = row.get('Name', '')
            fields['Number of Days in this cycle'] = int(no_of_actual_days)
            fields['Paid Days'] = int(row.get('No. of working days'))
            fields['1570 - Ticket Allowance'] = row.get('Ticket Allowance','')
            fields['1542 - OT'] = row.get('Fixed OT','')+row.get('Overtime', '')
            fields['1503 - Salary Diff Earn'] = row.get('Salary Diff Earn','')
            fields['1538 - Reimbursments'] = row.get('Reimbursments','')
            fields['1529 - KPI Incentive & Rewards'] = row.get('KPI Incentive & Rewards','')
            fields['2007 - Total Deduction'] = row.get('Salary Diff Ded(2007)','')+row.get('Salary Advance Ded.','')+row.get('Installments','')+row.get('Penalty','')+row.get('Total Deduction','')
            content_list.append(fields)
        df_output = pd.DataFrame(content_list).fillna('')

        leave_list = []
        df_output_leave = df_output[df_output['Number of Days in this cycle'] != df_output['Paid Days']]
        for index, row in df_output_leave.iterrows():
            end_date = self.calculate_leave_end_date(str(start_date), row['Number of Days in this cycle'] - row['Paid Days'])
            fields = {}
            fields['SAP ID'] = int(row.get('SAP ID', ''))
            fields['TM Name'] = row.get('TM Name', '')
            fields['Type of Leave '] = 'Unpaid Leave'
            fields['Start Date'] = start_date.strftime("%d/%m/%Y")
            fields['End Date'] = end_date.strftime("%d/%m/%Y")
            fields['Session'] = ''
            leave_list.append(fields)
        df_leave = pd.DataFrame(leave_list)
        self.segregate_client_file(df_output, df_leave)
    

    def segregate_client_file(self, df_output, df_leave):
        df_output1 = df_output[df_output['SAP ID'].astype(str).str.startswith('101')]
        df_leave1 = df_leave[df_leave['SAP ID'].astype(str).str.startswith('101')]
        output_filename = f'2101337_TK_0{self.today.month}_{self.today.year}_1000'
        self.modify_workbook(df_output1, df_leave1, output_filename)
        df_output2 = df_output[df_output['SAP ID'].astype(str).str.startswith('201')]
        df_leave2 = df_leave[df_leave['SAP ID'].astype(str).str.startswith('201')]
        output_filename = f'2100306_TK_0{self.today.month}_{self.today.year}_2000'         
        self.modify_workbook(df_output2, df_leave2, output_filename)
        try:
            shutil.move(self.input_filepath, os.path.join(self.BASE_DIR, 'Completed'))
        except Exception as e:
            print(f'Error - {e}')

    def modify_workbook(self, df_output, df_leave, output_filename):
        output_filepath = os.path.join(self.BASE_DIR, f'Output\{output_filename}.xlsx')
        with pd.ExcelWriter(output_filepath, engine="openpyxl") as writer:
            df_output = df_output.replace([0, 0.0], '')
            df_leave = df_leave.replace([0, 0.0], '')
            df_output.to_excel(writer, sheet_name="Input Sheet", index=False)
            df_leave.to_excel(writer, sheet_name="Leave", index=False)
        for sheet_name in ['Input Sheet', 'Leave']:
            wb = load_workbook(output_filepath)
            thick_border = Border(
                left=Side(border_style="medium", color="000000"),
                right=Side(border_style="medium", color="000000"),
                top=Side(border_style="medium", color="000000"),
                bottom=Side(border_style="medium", color="000000"),
            )

            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for row in ws.iter_rows():
                    for cell in row:
                        cell.border = thick_border   # Apply bold border to each cell

            ws = wb[sheet_name]
            fill = PatternFill(start_color="ED254E", end_color="ED254E", fill_type="solid")
            font = Font(color="FFFFFF", bold=True)  # white + bold
            alignment = Alignment(horizontal="center", vertical="center")
            # Apply style to header row
            for cell in ws[1]:
                cell.fill = fill
                cell.font = font
                cell.alignment = alignment

            ws.row_dimensions[1].height = 30 

            for col in ws.iter_cols(min_row=1, max_row=1):
                col_letter = col[0].column_letter
                ws.column_dimensions[col_letter].width = 20  # adjust as needed
            wb.save(output_filepath)

            
if __name__ == '__main__':
    obj = ConvertFile()
    obj.generate_outputfile()